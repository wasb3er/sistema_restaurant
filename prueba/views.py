from django.shortcuts import render, redirect, get_object_or_404
from django.http import JsonResponse, HttpResponse
from .models import Platillo, DetallePedido, Pedido, Empleado, Mesa, Categoria
from django.views.decorators.csrf import csrf_exempt, csrf_protect
from django.views.decorators.http import require_POST
from django.contrib.auth import authenticate, login
from django.contrib.auth.hashers import make_password, check_password
from django.contrib import messages
from functools import wraps
from django.contrib.sessions.backends.db import SessionStore
from django.utils.timezone import localtime
import json
from PIL import Image
from io import BytesIO
from django.core.files.base import ContentFile

import openpyxl
from openpyxl.styles import Font, Alignment
from django.http import HttpResponse
from django.utils import timezone
from datetime import timedelta
from django.db.models import Sum
from django.db.models.functions import TruncDate, TruncWeek

from django.http import JsonResponse
from django.views.decorators.csrf import csrf_exempt
from .models import Mesa, Pedido

from django.conf import settings
# from transbank.webpay.webpay_plus.transaction import Transaction

#Para la resolución de imagen
def redimensionar_imagen(imagen, ancho=612, alto=408):
    """Redimensiona cualquier imagen a 612x408 manteniendo calidad."""
    img = Image.open(imagen)

    # Convertir a RGB si viene en PNG con alpha
    if img.mode in ("RGBA", "P"):
        img = img.convert("RGB")

    img = img.resize((ancho, alto), Image.Resampling.LANCZOS)

    buffer = BytesIO()
    img.save(buffer, format="JPEG", quality=90)
    return ContentFile(buffer.getvalue())

# CAMBIO: se normaliza comparación de roles (minúsculas, sin espacios)
def rol_requerido(rol_permitido):
    """Valida que el usuario tenga sesión válida según el rol, con cookie independiente."""
    def decorador(vista_func):
        @wraps(vista_func)
        def _wrapped_view(request, *args, **kwargs):
            cookie_name = f"sessionid_{rol_permitido.strip().lower()}"
            session_key = request.COOKIES.get(cookie_name)

            if not session_key:
                return redirect('login_general')

            session = SessionStore(session_key=session_key)
            if not session or not session.get("empleado_rol"):
                return redirect('login_general')

            rol = session.get("empleado_rol", "").strip().lower()
            if rol != rol_permitido.strip().lower():
                return redirect('login_general')

            request.session = session
            return vista_func(request, *args, **kwargs)
        return _wrapped_view
    return decorador

# VISTAS PRINCIPALES
def index(request):
    return render(request, "index.html")


# CAMBIO: eliminado @login_required, solo se usa rol_requerido
@rol_requerido("Cocinero")
def cocinero(request):
    empleado_nombre = request.session.get("empleado_nombre", "Cocinero")
    return render(request, 'cocinero.html', {"empleado_nombre": empleado_nombre})


# def menu_view(request):
#     platillos = Platillo.objects.all()
#     return render(request, "menu.html", {"platillos": platillos, "user": request.user})
def menu(request):
    platillos = Platillo.objects.select_related("categoria").all()
    categorias = Categoria.objects.order_by("orden")
    return render(request, "menu.html", {
        "platillos": platillos,
        "categorias": categorias
    })



# CREAR PEDIDO
@csrf_exempt
@require_POST
def crear_pedido(request):
    try:
        data = json.loads(request.body)

        mesa_raw = data.get('mesa_id')
        if not mesa_raw:
            return JsonResponse({"success": False, "error": "ID de mesa no enviado"})

        try:
            mesa = Mesa.objects.get(numero=mesa_raw)
        except:
            return JsonResponse({"success": False, "error": "Mesa no encontrada"})

        nombre = data.get('nombre')
        personas = data.get('personas')
        platillos = data.get('platillos')

        if not platillos:
            return JsonResponse({'success': False, 'error': 'El carrito está vacío'})

        pedido = Pedido.objects.create(
            nombre_cliente=nombre,
            personas=personas,
            mesa=mesa,
            total=0,
            estado='nuevo'
        )

        total = 0
        for item in platillos:
            platillo = Platillo.objects.get(id=item['id'])
            cantidad = int(item['cantidad'])
            subtotal = platillo.precio * cantidad

            DetallePedido.objects.create(
                pedido=pedido,
                platillo=platillo,
                cantidad=cantidad,
                subtotal=subtotal
            )

            platillo.cantidad -= cantidad
            platillo.save()

            total += subtotal

        pedido.total = total
        pedido.save()

        mesa.estado = "ocupada"
        mesa.save()

        return JsonResponse({
            "success": True,
            "pedido_id": pedido.id,
            "mesa_id": mesa.id
        })

    except Exception as e:
        return JsonResponse({"success": False, "error": str(e)})

    

# CRUD DE PLATILLOS
@csrf_exempt
def crear_platillo(request):
    if request.method == 'POST':
        try:
            nombre = request.POST.get('nombre')
            descripcion = request.POST.get('descripcion')
            precio = request.POST.get('precio')
            cantidad = request.POST.get('cantidad', 0)
            imagen = request.FILES.get('imagen')

            nuevo_platillo = Platillo(
                nombre=nombre,
                descripcion=descripcion,
                precio=precio,
                cantidad=cantidad
            )
            # Si subieron una imagen: redimensionarla
            if imagen:
                imagen_redimensionada = redimensionar_imagen(imagen)
                nuevo_platillo.imagen.save(imagen.name, imagen_redimensionada, save=False)

            nuevo_platillo.save()
            return JsonResponse({'success': True})

        except Exception as e:
            return JsonResponse({'success': False, 'error': str(e)})

    return JsonResponse({'success': False, 'error': 'Método no permitido'}, status=405)



def listar_platillos(request):
    platillos = list(Platillo.objects.values())
    return JsonResponse({'platillos': platillos})


@csrf_exempt
def editar_platillo(request, id):
    try:
        if request.method != "POST":
            return JsonResponse({"success": False, "error": "Método no permitido"}, status=405)

        platillo = get_object_or_404(Platillo, id=id)

        # Actualizar campos
        platillo.nombre = request.POST.get("nombre", platillo.nombre)
        platillo.descripcion = request.POST.get("descripcion", platillo.descripcion)

        # Convertir precio y cantidad (evita errores si vienen vacíos)
        precio = request.POST.get("precio")
        cantidad = request.POST.get("cantidad")

        if precio:
            platillo.precio = float(precio)

        if cantidad:
            platillo.cantidad = int(cantidad)

        # Si llega una imagen nueva, reemplazarla
        if "imagen" in request.FILES:
            imagen_original = request.FILES["imagen"]
            imagen_redimensionada = redimensionar_imagen(imagen_original)
            platillo.imagen.save(imagen_original.name, imagen_redimensionada, save=False)

        platillo.save()

        return JsonResponse({"success": True})

    except Exception as e:
        return JsonResponse({"success": False, "error": str(e)}, status=500)




@csrf_exempt
def eliminar_platillo(request, id):
    if request.method == 'DELETE':
        try:
            platillo = get_object_or_404(Platillo, id=id)
            platillo.delete()
            return JsonResponse({'success': True})
        except Exception as e:
            return JsonResponse({'success': False, 'error': str(e)})
    return JsonResponse({'success': False, 'error': 'Método no permitido'}, status=405)


# CAMBIO: eliminado @login_required
@rol_requerido("administrador")
def admin_menu(request):
    platillos = Platillo.objects.all()
    empleados = Empleado.objects.all()
    pedidos = Pedido.objects.all()
    return render(request, 'admin/admin-menu.html', {
        'platillos': platillos,
        'empleados': empleados,
        'pedidos': pedidos,
    })

@rol_requerido("administrador")
def admin_reportes(request):
    return render(request, "admin/reportes.html")

# CAMBIO: eliminado @login_required
@rol_requerido("administrador")
def cambiar_estado_pedido(request, pedido_id):
    try:
        pedido = Pedido.objects.get(id=pedido_id)
        pedido.estado = 'terminado' if pedido.estado == 'pendiente' else 'pendiente'
        pedido.save()
        return JsonResponse({'success': True, 'nuevo_estado': pedido.estado})
    except Pedido.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'Pedido no encontrado'}, status=404)


# CAMBIO: eliminado @login_required
@rol_requerido("administrador")
def eliminar_pedido(request, pedido_id):
    pedido = get_object_or_404(Pedido, id=pedido_id)
    pedido.delete()
    return redirect('admin_menu')


# CAMBIO: eliminado @login_required
@rol_requerido("administrador")
def pedidos_pendientes(request):
    pedidos = Pedido.objects.filter(estado='pendiente').order_by('-id')
    data = [
        {
            "id": p.id,
            "nombre_cliente": p.nombre_cliente,
            "personas": p.personas,
            "total": float(p.total),
            "fecha": p.fecha.strftime("%Y-%m-%d %H:%M:%S"),
        }
        for p in pedidos
    ]
    return JsonResponse(data, safe=False)


# LOGIN GENERAL
@csrf_protect
def login_general(request):
    if request.method == 'POST':
        tipo = request.POST.get("tipoLogin")
        username = request.POST.get("username")
        password = request.POST.get("password")

#LOGIN ADMIN
        if tipo == "admin":
            user = authenticate(username=username, password=password)
            if user:
                session = SessionStore()
                session["empleado_id"] = user.id
                session["empleado_nombre"] = user.username
                session["empleado_rol"] = "administrador"
                session.save()

                response = redirect('admin_menu')
                response.set_cookie("sessionid_administrador", session.session_key, httponly=True)
                return response
            else:
                messages.error(request, "Credenciales inválidas para administrador.")
                return render(request, 'login.html')

# LOGIN EMPLEADO
        elif tipo == "empleado":
            empleado = Empleado.objects.filter(username=username).first()

            if not empleado:
                messages.error(request, "Usuario no encontrado.")
                return render(request, 'login.html')

            # Validar contraseña
            if not check_password(password, empleado.password):
                messages.error(request, "Contraseña incorrecta.")
                return render(request, 'login.html')

            # Validación de estado laboral
            if empleado.estado_laboral != "trabajando":
                messages.error(request, "Tu cuenta está inhabilitada temporalmente. Contacta al administrador.")
                return render(request, 'login.html')

            # Validación de activo
            if not empleado.activo:
                messages.error(request, "Tu cuenta está inactiva. Contacta al administrador.")
                return render(request, 'login.html')

            # Si pasa todas las validaciones: crear sesión
            session = SessionStore()
            session["empleado_id"] = empleado.id
            session["empleado_nombre"] = empleado.nombre
            session["empleado_rol"] = empleado.rol
            session.save()

            rol = empleado.rol.strip().lower()
            if rol == "mesero":
                response = redirect('mesero_menu')
                cookie_name = "sessionid_mesero"
            elif rol == "cocinero":
                response = redirect('cocina_menu')
                cookie_name = "sessionid_cocinero"
            else:
                response = redirect('admin_menu')
                cookie_name = "sessionid_admin"

            response.set_cookie(cookie_name, session.session_key, httponly=True)
            return response
        
        else:
            messages.error(request, "Selecciona el tipo de usuario antes de ingresar.")
            return render(request, 'login.html')

    # ⬇⬇⬇ AQUI ESTABA TU ERROR — ESTE RETURN FALTABA ⬇⬇⬇
    return render(request, 'login.html')

#LOGOUT POR ROL
def logout_por_rol(request, rol):
    cookie_name = f"sessionid_{rol.strip().lower()}"
    response = redirect('login_general')
    response.delete_cookie(cookie_name)
    return response

#eliminado @login_required
@rol_requerido("cocinero")
def cocina_menu(request):
    empleado_nombre = request.session.get("empleado_nombre", "Cocinero")
    return render(request, 'cocinero.html', {"empleado_nombre": empleado_nombre})


#eliminado @login_required
@rol_requerido("mesero")
def mesero_menu(request):
    empleado_nombre = request.session.get("empleado_nombre", "Mesero")
    return render(request, 'mesero_menu.html', {"empleado_nombre": empleado_nombre})


#eliminado @login_required
@csrf_exempt
@rol_requerido("administrador")
def crear_empleado(request):
    if request.method == "POST":
        nombre = request.POST.get("nombre")
        username = request.POST.get("username")
        password = request.POST.get("password")
        rol = request.POST.get("rol", "Mesero")

        if not nombre or not username or not password:
            return JsonResponse({"success": False, "error": "Faltan datos obligatorios"})

        if Empleado.objects.filter(username=username).exists():
            return JsonResponse({"success": False, "error": "El usuario ya existe"})

        nuevo = Empleado(
            nombre=nombre,
            username=username,
            rol=rol,
            password=make_password(password),
            activo=True
        )
        nuevo.save()

        return JsonResponse({
            "success": True,
            "message": f"Empleado '{nombre}' agregado correctamente."
        })

    return JsonResponse({"success": False, "error": "Método no permitido"}, status=405)



from django.utils.timezone import localtime

#API para mostrar pedidos NUEVOS (vista del MESERO)
def pedidos_nuevos_api(request):
    try:
        pedidos = Pedido.objects.filter(estado__in=['nuevo', 'pendiente']).order_by('-fecha')
        data = [
            {
                "id": p.id,
                "nombre_cliente": p.nombre_cliente,
                "personas": p.personas,
                "total": float(p.total),
                "fecha": p.fecha.strftime("%Y-%m-%d %H:%M:%S") if p.fecha else None,
                "estado": p.estado
            }
            for p in pedidos
        ]
        return JsonResponse({"pedidos": data})
    except Exception as e:
        return JsonResponse({"error": str(e)}, status=500)


#API para que el MESERO envíe un pedido a cocina
@csrf_exempt
def enviar_a_cocina(request, pedido_id):
    try:
        pedido = Pedido.objects.get(pk=pedido_id)

        # Solo se puede enviar a cocina si está en nuevo o pendiente
        if pedido.estado not in ['nuevo', 'pendiente']:
            return JsonResponse({
                "success": False,
                "error": "El pedido no puede enviarse a cocina desde su estado actual."
            })

        pedido.estado = 'enviado_a_cocina'
        pedido.save()

        return JsonResponse({
            "success": True,
            "message": "Pedido enviado a cocina correctamente."
        })

    except Pedido.DoesNotExist:
        return JsonResponse({
            "success": False,
            "error": "Pedido no encontrado."
        })

    except Exception as e:
        return JsonResponse({
            "success": False,
            "error": str(e)
        })

    

#API para pedidos del cocinero (en cocina o listos)
def pedidos_en_cocina_api(request):
    try:
        # Pedidos que el cocinero debe ver
        pedidos = (
            Pedido.objects
            .filter(estado__in=["enviado_a_cocina", "listo"])
            .order_by("-fecha")
        )

        data = []
        for p in pedidos:
            detalles = (
                DetallePedido.objects
                .filter(pedido=p)
                .select_related("platillo")
            )

            items = []
            for d in detalles:
                items.append(
                    {
                        "nombre": d.platillo.nombre,
                        "cantidad": d.cantidad,
                        "subtotal": float(d.subtotal),
                    }
                )

            data.append(
                {
                    "id": p.id,
                    "nombre_cliente": p.nombre_cliente,
                    "personas": p.personas,
                    "total": float(p.total),
                    "fecha": p.fecha.strftime("%Y-%m-%d %H:%M:%S") if p.fecha else None,
                    "estado": p.estado,
                    "items": items,
                }
            )

        return JsonResponse({"pedidos": data})

    except Exception as e:
        return JsonResponse({"error": str(e)}, status=500)


#Marcar pedido como listo
@csrf_exempt
def marcar_listo(request, pedido_id):
    try:
        pedido = Pedido.objects.get(pk=pedido_id)

        # Validar transición correcta del flujo
        if pedido.estado != "enviado_a_cocina":
            return JsonResponse(
                {
                    "success": False,
                    "error": "Solo se pueden marcar como 'listo' los pedidos enviados a cocina.",
                }
            )

        pedido.estado = "listo"
        pedido.save()

        return JsonResponse({"success": True})

    except Pedido.DoesNotExist:
        return JsonResponse(
            {
                "success": False,
                "error": "Pedido no encontrado.",
            }
        )
    except Exception as e:
        return JsonResponse(
            {
                "success": False,
                "error": str(e),
            }
        )



#Volver a estado "enviado_a_cocina"
@csrf_exempt
def volver_cocina(request, pedido_id):
    try:
        pedido = Pedido.objects.get(pk=pedido_id)
        pedido.estado = 'enviado_a_cocina'
        pedido.save()
        return JsonResponse({"success": True})
    except Pedido.DoesNotExist:
        return JsonResponse({"success": False, "error": "Pedido no encontrado."})
    except Exception as e:
        return JsonResponse({"success": False, "error": str(e)})

    
#API para pedidos listos (mesero los ve para entregar)
def pedidos_listos_api(request):
    try:
        pedidos = Pedido.objects.filter(estado='listo').order_by('-fecha')
        data = [
            {
                "id": p.id,
                "nombre_cliente": p.nombre_cliente,
                "personas": p.personas,
                "total": float(p.total),
                "fecha": p.fecha.strftime("%Y-%m-%d %H:%M:%S") if p.fecha else None,
                "estado": p.estado
            }
            for p in pedidos
        ]
        return JsonResponse({"pedidos": data})
    except Exception as e:
        return JsonResponse({"error": str(e)}, status=500)

@csrf_exempt
def marcar_entregado(request, pedido_id):
    try:
        pedido = Pedido.objects.get(pk=pedido_id)

        if pedido.estado != "listo":
            return JsonResponse(
                {
                    "success": False,
                    "error": "Solo se pueden marcar como entregados los pedidos en estado 'listo'.",
                }
            )

        pedido.estado = "entregado"
        pedido.save()

        return JsonResponse({"success": True})

    except Pedido.DoesNotExist:
        return JsonResponse(
            {
                "success": False,
                "error": "Pedido no encontrado.",
            }
        )
    except Exception as e:
        return JsonResponse(
            {
                "success": False,
                "error": str(e),
            }
        )


@csrf_exempt
def liberar_mesa(request, pedido_id):
    try:
        pedido = Pedido.objects.select_related("mesa").get(pk=pedido_id)
        mesa = pedido.mesa

        if not mesa:
            return JsonResponse(
                {
                    "success": False,
                    "error": "El pedido no tiene mesa asociada.",
                }
            )

        # Solo debería liberarse cuando el pedido ya está entregado
        if pedido.estado != "entregado":
            return JsonResponse(
                {
                    "success": False,
                    "error": "Solo se pueden liberar mesas con pedidos entregados.",
                }
            )

        # Marcar la mesa como disponible y cortar vínculo
        mesa.estado = "disponible"
        mesa.updated_at = timezone.now()
        mesa.save()

        pedido.estado = "terminado"
        pedido.mesa = None
        pedido.save()

        return JsonResponse({"success": True})

    except Pedido.DoesNotExist:
        return JsonResponse(
            {
                "success": False,
                "error": "Pedido no encontrado.",
            }
        )
    except Exception as e:
        return JsonResponse(
            {
                "success": False,
                "error": str(e),
            }
        )


    

# API MESAS: ESTADO VISUAL PARA EL MESERO
def mesas_api(request):
    mesas = Mesa.objects.all().order_by("numero")
    
    data = []
    for m in mesas:

        # Buscar pedido activo de la mesa
        pedido = (
            Pedido.objects.filter(
                mesa=m,
                estado__in=[
                    "nuevo",
                    "pendiente",
                    "enviado_a_cocina",
                    "listo",
                    "entregado",
                ]
            )
            .order_by("-id")
            .first()
        )

        if pedido:
            pedido_json = {
                "id": pedido.id,
                "estado": pedido.estado,
                "total": float(pedido.total),
            }
            estado_final = "ocupada"
        else:
            pedido_json = None
            estado_final = "disponible"

        data.append({
            "id": m.id,
            "numero": m.numero,
            "estado": estado_final,
            "pedido": pedido_json,
        })

    return JsonResponse({"mesas": data})




def pedido_por_mesa(request, mesa_id):
    
    try:
        mesa = Mesa.objects.get(pk=mesa_id)
        pedido = Pedido.objects.filter(
            mesa=mesa
        ).exclude(
            estado="terminado"
        ).first()

        if not pedido:
            return JsonResponse({"pedido": None})

        return JsonResponse({
            "pedido": {
                "id": pedido.id,
                "estado": pedido.estado
            }
        })

    except Mesa.DoesNotExist:
        return JsonResponse({"error": "Mesa no encontrada"}, status=404)


def api_pedidos(request):
    try:
        pedidos = Pedido.objects.order_by('-id')

        data = []

        for p in pedidos:
            # Fecha segura
            fecha = None
            if hasattr(p, "fecha") and p.fecha:
                try:
                    fecha = p.fecha.strftime("%Y-%m-%d %H:%M:%S")
                except:
                    fecha = None

            data.append({
                "id": p.id,
                "cliente": getattr(p, "nombre_cliente", "N/A") or "N/A",
                "personas": getattr(p, "personas", 0) or 0,
                "total": float(getattr(p, "total", 0) or 0),
                "estado": getattr(p, "estado", "N/A"),
                "fecha": fecha,
            })

        return JsonResponse({"pedidos": data})

    except Exception as e:
        return JsonResponse({"error": str(e)}, status=500)


def api_platillos(request):
    platillos = Platillo.objects.select_related("categoria").all()

    data = []
    for p in platillos:
        data.append({
            "id": p.id,
            "nombre": p.nombre,
            "descripcion": p.descripcion,
            "precio": float(p.precio),
            "cantidad": p.cantidad,
            # NUEVO: info de categoría
            "categoria_id": p.categoria.id if p.categoria else None,
            "categoria_slug": p.categoria.slug if p.categoria else None,
            # opcional: por si quieres usarlo después
            "es_del_dia": p.es_del_dia,
        })

    return JsonResponse(data, safe=False)

@csrf_exempt
def toggle_empleado_activo(request, empleado_id):
    try:
        empleado = get_object_or_404(Empleado, id=empleado_id)

        # Invertir el estado
        empleado.activo = not empleado.activo
        empleado.save()

        return JsonResponse({
            "success": True,
            "activo": empleado.activo,
            "message": "Empleado actualizado correctamente"
        })
    except Exception as e:
        return JsonResponse({"success": False, "error": str(e)})
    
@csrf_exempt
def cambiar_estado_empleado(request, empleado_id):
    try:
        if request.method != "POST":
            return JsonResponse({"success": False, "error": "Método no permitido"}, status=405)

        data = json.loads(request.body.decode("utf-8"))
        nuevo_estado = data.get("estado")

        if nuevo_estado not in ["trabajando", "vacaciones", "licencia", "despedido"]:
            return JsonResponse({"success": False, "error": "Estado no válido"}, status=400)

        empleado = get_object_or_404(Empleado, id=empleado_id)
        empleado.estado_laboral = nuevo_estado

        # Opcional: si está despedido → no puede iniciar sesión
        if nuevo_estado == "trabajando":
            empleado.activo = True
        else:
            empleado.activo = False

        empleado.save()

        return JsonResponse({"success": True, "estado": nuevo_estado})

    except Exception as e:
        return JsonResponse({"success": False, "error": str(e)}, status=500)



@rol_requerido("administrador")
def reporte_ventas_excel(request):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Ventas"

    # Encabezados
    headers = ["ID Pedido", "Cliente", "Personas", "Total", "Estado", "Fecha"]
    ws.append(headers)

    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")

    pedidos = Pedido.objects.order_by("-fecha")

    for p in pedidos:
        ws.append([
            p.id,
            p.nombre_cliente,
            p.personas,
            float(p.total),
            p.estado,
            p.fecha.strftime("%Y-%m-%d %H:%M")
        ])

    # Preparar respuesta
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response['Content-Disposition'] = 'attachment; filename="reporte_ventas.xlsx"'

    wb.save(response)
    return response

@rol_requerido("administrador")
def reporte_platillos_excel(request):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Platillos"

    headers = ["Platillo", "Total Vendido", "Ingresos Generados"]
    ws.append(headers)

    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")

    platillos = Platillo.objects.all()

    for pl in platillos:
        detalles = DetallePedido.objects.filter(platillo=pl)
        total_vendido = sum(d.cantidad for d in detalles)
        ingresos = sum(float(d.subtotal) for d in detalles)

        ws.append([pl.nombre, total_vendido, ingresos])

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response['Content-Disposition'] = 'attachment; filename="reporte_platillos.xlsx"'

    wb.save(response)
    return response

@csrf_exempt
@rol_requerido("administrador")
def api_reportes_filtrar(request):
    try:
        # ============================================
        # MÉTODO NO PERMITIDO
        # ============================================
        if request.method != "POST":
            return JsonResponse({"success": False, "error": "Método no permitido"}, status=405)

        # ============================================
        # OBTENER BODY
        # ============================================
        try:
            body = json.loads(request.body.decode("utf-8"))
        except:
            body = {}

        rango = body.get("rango", "hoy")
        cliente = (body.get("cliente") or "").strip()
        agrupacion = body.get("agrupacion", "dia")

        # ============================================
        # FECHA LOCAL CHILE
        # ============================================
        ahora_local = localtime(timezone.now())

        def inicio_dia(dt):
            dt = localtime(dt)
            return dt.replace(hour=0, minute=0, second=0, microsecond=0)

        inicio = fin = None

        if rango == "hoy":
            inicio = inicio_dia(ahora_local)
            fin = ahora_local

        elif rango == "ayer":
            ay = inicio_dia(ahora_local - timedelta(days=1))
            inicio = ay
            fin = ay + timedelta(days=1)

        elif rango == "semana":
            lunes = inicio_dia(ahora_local - timedelta(days=ahora_local.weekday()))
            inicio = lunes
            fin = ahora_local

        elif rango == "mes":
            primero = inicio_dia(ahora_local.replace(day=1))
            inicio = primero
            fin = ahora_local

        elif rango == "todos":
            inicio = None
            fin = None

        # ============================================
        # FILTRAR PEDIDOS
        # ============================================
        pedidos = Pedido.objects.all()

        if inicio:
            pedidos = pedidos.filter(fecha__gte=inicio)
        if fin:
            pedidos = pedidos.filter(fecha__lte=fin)

        if cliente:
            pedidos = pedidos.filter(nombre_cliente__icontains=cliente)

        pedidos = pedidos.order_by("-fecha")

        # ============================================
        # TABLA DETALLADA
        # ============================================
        pedidos_data = []
        for p in pedidos:
            try:
                fecha_str = localtime(p.fecha).strftime("%Y-%m-%d %H:%M")
            except:
                fecha_str = ""

            pedidos_data.append({
                "id": p.id,
                "cliente": p.nombre_cliente or "N/A",
                "fecha": fecha_str,
                "total": float(p.total or 0),
                "estado": p.estado,
            })

        # ============================================
        # KPIs
        # ============================================
        cantidad = pedidos.count()
        total_vendido = sum(float(p.total or 0) for p in pedidos)
        ticket_promedio = total_vendido / cantidad if cantidad > 0 else 0

        # ============================================
        # TOP PLATILLOS
        # ============================================
        detalles = DetallePedido.objects.filter(pedido__in=pedidos).select_related("platillo")

        stats = {}
        top_platillos = []

        for d in detalles:
            if not d.platillo:
                continue

            nombre = d.platillo.nombre

            if agrupacion == "platillo_ingresos":
                valor = float(d.subtotal or 0)
            else:
                valor = d.cantidad or 0

            stats[nombre] = stats.get(nombre, 0) + valor

        platillo_top = None
        if stats:
            ordenados = sorted(stats.items(), key=lambda x: x[1], reverse=True)
            top_platillos = [{"nombre": n, "valor": v} for n, v in ordenados]
            platillo_top = ordenados[0][0]

        # ============================================
        # NUEVO: SERIE TEMPORAL DE TOTAL VENDIDO
        # ============================================
        from django.db.models.functions import TruncHour, TruncDay, TruncMonth
        from django.db.models import Sum

        grafico_total = []

        if agrupacion == "total_vendido":

            # HOY / AYER → POR HORA
            if rango in ["hoy", "ayer"]:
                datos = (
                    pedidos
                    .annotate(periodo=TruncHour("fecha"))
                    .values("periodo")
                    .annotate(total=Sum("total"))
                    .order_by("periodo")
                )
                grafico_total = [
                    {"etiqueta": d["periodo"].strftime("%H:%M"), "total": float(d["total"])}
                    for d in datos
                ]

            # SEMANA / MES → POR DÍA
            elif rango in ["semana", "mes"]:
                datos = (
                    pedidos
                    .annotate(periodo=TruncDay("fecha"))
                    .values("periodo")
                    .annotate(total=Sum("total"))
                    .order_by("periodo")
                )
                grafico_total = [
                    {"etiqueta": d["periodo"].strftime("%Y-%m-%d"), "total": float(d["total"])}
                    for d in datos
                ]

            # TODOS → POR MES
            elif rango == "todos":
                datos = (
                    pedidos
                    .annotate(periodo=TruncMonth("fecha"))
                    .values("periodo")
                    .annotate(total=Sum("total"))
                    .order_by("periodo")
                )
                grafico_total = [
                    {"etiqueta": d["periodo"].strftime("%Y-%m"), "total": float(d["total"])}
                    for d in datos
                ]

        # ============================================
        # RESPUESTA FINAL
        # ============================================
        return JsonResponse({
            "success": True,
            "kpis": {
                "total_vendido": round(total_vendido, 2),
                "cantidad_pedidos": cantidad,
                "ticket_promedio": round(ticket_promedio, 2),
                "platillo_top": platillo_top,
            },
            "pedidos": pedidos_data,
            "top_platillos": top_platillos,
            "grafico_total": grafico_total,   # << NUEVO
        })

    except Exception as e:
        return JsonResponse(
            {"success": False, "error": f"Error interno en reportes: {str(e)}"},
            status=500
        )







@csrf_exempt
@rol_requerido("administrador")
def reporte_ventas_filtrado_excel(request):
    """
    Exporta un archivo Excel con los mismos filtros usados por api_reportes_filtrar:
    - rango: hoy, ayer, semana, mes
    - cliente: texto parcial
    - agrupacion: no afecta la tabla, pero se incluye en cabecera
    """
    try:
        rango = request.GET.get("rango", "hoy")
        cliente = (request.GET.get("cliente") or "").strip()
        agrupacion = request.GET.get("agrupacion", "dia")
    except:
        rango = "hoy"
        cliente = ""
        agrupacion = "dia"

    # ==========================================================
    # CALCULAR RANGO DE FECHAS
    # ==========================================================
    ahora = localtime(timezone.now())

    def inicio_dia(dt):
        d = localtime(dt)
        return d.replace(hour=0, minute=0, second=0, microsecond=0)

    inicio = fin = None

    if rango == "hoy":
        inicio = inicio_dia(ahora)
        fin = ahora

    elif rango == "ayer":
        ayer = inicio_dia(ahora - timedelta(days=1))
        inicio = ayer
        fin = ayer + timedelta(days=1)

    elif rango == "semana":
        lunes = inicio_dia(ahora - timedelta(days=ahora.weekday()))
        inicio = lunes
        fin = ahora

    elif rango == "mes":
        primero = inicio_dia(ahora.replace(day=1))
        inicio = primero
        fin = ahora

    # ==========================================================
    # FILTRAR PEDIDOS
    # ==========================================================
    pedidos_qs = Pedido.objects.all()

    if inicio:
        pedidos_qs = pedidos_qs.filter(fecha__gte=inicio)
    if fin:
        pedidos_qs = pedidos_qs.filter(fecha__lte=fin)

    if cliente:
        pedidos_qs = pedidos_qs.filter(nombre_cliente__icontains=cliente)

    pedidos_qs = pedidos_qs.order_by("-fecha")

    # ==========================================================
    # CREAR EXCEL
    # ==========================================================
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Reporte de Ventas"

    # Título Headers
    ws["A1"] = "Reporte de Ventas Filtrado"
    ws["A1"].font = Font(size=14, bold=True)

    ws["A3"] = f"Rango aplicado: {rango}"
    ws["A4"] = f"Cliente: {cliente if cliente else 'Todos'}"
    ws["A5"] = f"Agrupación: {agrupacion}"

    # Encabezado tabla
    headers = ["ID Pedido", "Cliente", "Personas", "Fecha", "Total", "Estado"]
    ws.append([])  # espacio
    ws.append(headers)

    # Estilos de encabezado
    for col in range(1, len(headers) + 1):
        cell = ws.cell(row=7, column=col)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")

    # Agregar datos
    for pedido in pedidos_qs:
        try:
            fecha_str = localtime(pedido.fecha).strftime("%Y-%m-%d %H:%M")
        except:
            fecha_str = ""

        ws.append([
            pedido.id,
            pedido.nombre_cliente,
            pedido.personas,
            fecha_str,
            float(pedido.total or 0),
            pedido.estado
        ])

    # ==========================================================
    # RESPUESTA HTTP DEL ARCHIVO
    # ==========================================================
    nombre_archivo = "reporte_filtrado.xlsx"

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = f'attachment; filename="{nombre_archivo}"'
    wb.save(response)

    return response



def api_categorias(request):
    categorias = Categoria.objects.all().order_by("orden")
    data = []

    for c in categorias:
        data.append({
            "id": c.id,
            "nombre": c.nombre,
            "slug": c.slug,
            "orden": c.orden
        })

    return JsonResponse({"categorias": data})


from django.http import JsonResponse
from django.views.decorators.csrf import csrf_exempt

@csrf_exempt
def api_pedido_detalle(request, pedido_id):
    try:
        pedido = Pedido.objects.get(id=pedido_id)

        detalles = DetallePedido.objects.filter(pedido=pedido).select_related("platillo")

        detalle_items = []
        for d in detalles:
            detalle_items.append({
                "platillo": d.platillo.nombre,
                "cantidad": d.cantidad,
                "subtotal": float(d.subtotal),
            })

        return JsonResponse({
            "success": True,
            "pedido": {
                "id": pedido.id,
                "cliente": pedido.nombre_cliente,
                "personas": pedido.personas,
                "fecha": pedido.fecha.strftime("%Y-%m-%d %H:%M"),
                "total": float(pedido.total),
                "estado": pedido.estado,
                "mesa": pedido.mesa.numero if pedido.mesa else None,
                "items": detalle_items
            }
        })

    except Pedido.DoesNotExist:
        return JsonResponse({"success": False, "error": "Pedido no encontrado"}, status=404)

@csrf_exempt
def agregar_item_pedido(request, pedido_id):
    try:
        data = json.loads(request.body)

        platillo_id = data.get("platillo_id")
        cantidad = int(data.get("cantidad", 1))

        pedido = Pedido.objects.get(id=pedido_id)
        platillo = Platillo.objects.get(id=platillo_id)

        subtotal = platillo.precio * cantidad

        DetallePedido.objects.create(
            pedido=pedido,
            platillo=platillo,
            cantidad=cantidad,
            subtotal=subtotal
        )

        pedido.total += subtotal
        pedido.save()

        return JsonResponse({"success": True})
    except Exception as e:
        return JsonResponse({"success": False, "error": str(e)})

#Asignar mesero
# @csrf_exempt
# def asignar_mesero(request, mesa_id):
#     try:
#         data = json.loads(request.body)
#         mesero_id = data.get("mesero_id")

#         mesa = Mesa.objects.get(pk=mesa_id)
#         mesero = Empleado.objects.get(pk=mesero_id, rol="Mesero")

#         mesa.mesero = mesero
#         mesa.save()

#         return JsonResponse({"success": True})
#     except Exception as e:
#         return JsonResponse({"success": False, "error": str(e)}, status=500)


import json
from django.http import JsonResponse
from django.shortcuts import render
from django.views.decorators.csrf import csrf_exempt

# IMPORT CORRECTO PARA SDK 4.0.0
from transbank.webpay.webpay_plus.transaction import Transaction


# ================================================================
# GUARDAR CARRITO
# ================================================================
@csrf_exempt
def guardar_carrito(request):
    data = json.loads(request.body)

    carrito = data.get("carrito", [])
    cliente = data.get("cliente")
    personas = data.get("personas")
    mesa_id = data.get("mesa_id")

    total = sum(item["precio"] * item["cantidad"] for item in carrito)

    request.session["carrito"] = carrito
    request.session["cliente"] = cliente
    request.session["personas"] = personas
    request.session["mesa_id"] = mesa_id
    request.session["pago_total"] = total
    request.session.save()

    return JsonResponse({"success": True})


# ================================================================
# RESUMEN DEL PAGO
# ================================================================
def pago_resumen(request):
    carrito = request.session.get("carrito", [])
    total = request.session.get("pago_total", 0)

    return render(request, "pago/pago_resumen.html", {
        "carrito": carrito,
        "total": total
    })


# ================================================================
# INICIAR TRANSACCIÓN WEBPAY
# ================================================================
@csrf_exempt
def webpay_iniciar(request):

    total = request.session.get("pago_total")

    if not total:
        return JsonResponse({"error": "Total inválido"}, status=400)

    if not request.session.session_key:
        request.session.save()

    buy_order = f"ORD{request.session.session_key[-6:]}"
    session_id = request.session.session_key

    try:
        # Transaction directa, sin options, sin integration_type
        tx = Transaction()

        response = tx.create(
            buy_order=buy_order,
            session_id=session_id,
            amount=total,
            return_url="http://localhost:8000/pago/webpay/confirmar/"
        )

        # Guardar sesión
        request.session["buy_order"] = buy_order
        request.session["total_pagado"] = total
        request.session.save()

        return JsonResponse({
            "url": response["url"],
            "token": response["token"]
        })

    except Exception as e:
        return JsonResponse({
            "error": "Error iniciando pago",
            "detalle": str(e)
        }, status=500)


# ================================================================
# CONFIRMACIÓN
# ================================================================
@csrf_exempt
def webpay_retorno(request):
    token = request.GET.get("token_ws")

    if not token:
        return render(request, "pago/pago_error.html", {"status": "TOKEN_MISSING"})

    # =====================================================
    # A) *** EVITAR PROCESAR EL MISMO TOKEN 2 VECES ***
    #    (Por ahora usamos sesión; luego agregaremos tabla PagoWebpay)
    # =====================================================
    processed_tokens = request.session.get("processed_tokens", [])

    if token in processed_tokens:
        # Ya se procesó este pago → NO crear pedido de nuevo
        return render(request, "pago/pago_exitoso.html", {
            "buy_order": request.session.get("buy_order"),
            "amount": request.session.get("pago_total"),
            "mesa": request.session.get("mesa_id"),
        })

    try:
        tx = Transaction()
        response = tx.commit(token)

        # =====================================================
        # B) Pago autorizado correctamente
        # =====================================================
        if response["status"] == "AUTHORIZED":

            carrito = request.session.get("carrito", [])
            cliente = request.session.get("cliente")
            personas = request.session.get("personas")
            mesa_raw = request.session.get("mesa_id")
            total = request.session.get("pago_total")
            buy_order = request.session.get("buy_order")

            # =====================================================
            # C) Obtener mesa correctamente
            # =====================================================
            mesa = None
            if mesa_raw is not None:
                try:
                    mesa_int = int(mesa_raw)
                    mesa = Mesa.objects.filter(id=mesa_int).first() or \
                           Mesa.objects.filter(numero=mesa_int).first()
                except:
                    mesa = Mesa.objects.filter(numero=mesa_raw).first()

            # =====================================================
            # D) Crear pedido (SOLO 1 VEZ)
            # =====================================================
            pedido = Pedido.objects.create(
                nombre_cliente=cliente,
                personas=personas,
                mesa=mesa,
                total=total,
                estado="pendiente"
            )

            # =====================================================
            # E) Crear detalles del pedido
            # =====================================================
            for item in carrito:
                platillo = Platillo.objects.get(id=item["id"])
                cantidad = int(item["cantidad"])
                subtotal = float(platillo.precio) * cantidad

                DetallePedido.objects.create(
                    pedido=pedido,
                    platillo=platillo,
                    cantidad=cantidad,
                    subtotal=subtotal
                )

                # Descontar stock
                platillo.cantidad -= cantidad
                platillo.save()

            # =====================================================
            # F) Marcar mesa como ocupada
            # =====================================================
            if mesa:
                from django.utils import timezone
                mesa.estado = "ocupada"
                mesa.updated_at = timezone.now()
                mesa.save()

            # =====================================================
            # G) MARCAR TOKEN COMO PROCESADO (evita duplicación)
            # =====================================================
            processed_tokens.append(token)
            request.session["processed_tokens"] = processed_tokens

            # =====================================================
            # H) LIMPIAR datos del carrito
            # =====================================================
            request.session["carrito"] = []
            request.session["pago_total"] = 0
            request.session["mesa_id"] = None
            request.session.save()

            # =====================================================
            # I) Mostrar pantalla final
            # =====================================================
            return render(request, "pago/pago_exitoso.html", {
                "buy_order": buy_order,
                "amount": total,
                "pedido_id": pedido.id,
                "mesa": mesa.numero if mesa else None,
            })

        # =====================================================
        # Pago rechazado
        # =====================================================
        return render(request, "pago/pago_error.html", {"status": response["status"]})

    except Exception as e:
        return render(request, "pago/pago_error.html", {"status": str(e)})



from django.shortcuts import render, get_object_or_404
from django.views.decorators.csrf import csrf_protect
from django.urls import reverse
from django.utils import timezone

@csrf_protect
def mesa_qr(request):
    """
    Flujo cliente:
    1) GET → formulario inicial
    2) POST vista=form → mostrar selección de mesas
    3) POST vista=mesa → mostrar formulario de PIN
    4) POST vista=pin → validar y continuar
    """

    # ====== GET: FORMULARIO ======
    if request.method == "GET":
        return render(request, "cliente/mesa_qr.html", {
            "step": "form",
            "errores": [],
            "nombre": "",
            "personas": "",
        })

    # ====== POST ======
    vista = request.POST.get("vista")

    # -------------------------------
    # PASO 1 → PASO 2
    # -------------------------------
    if vista == "form":

        nombre = (request.POST.get("nombre") or "").strip()
        personas = (request.POST.get("personas") or "").strip()
        errores = []

        if not nombre:
            errores.append("Debes ingresar tu nombre.")
        if not personas.isdigit() or int(personas) <= 0:
            errores.append("Cantidad de personas inválida.")

        if errores:
            return render(request, "cliente/mesa_qr.html", {
                "step": "form",
                "errores": errores,
                "nombre": nombre,
                "personas": personas,
            })

        return render(request, "cliente/mesa_qr.html", {
            "step": "mesas",
            "errores": [],
            "nombre": nombre,
            "personas": personas,
        })

    # -------------------------------
    # PASO 2 → PASO 3 (PIN)
    # -------------------------------
    if vista == "mesa":

        mesa_numero = request.POST.get("mesa_numero")
        nombre = request.POST.get("nombre")
        personas = request.POST.get("personas")

        mesa = get_object_or_404(Mesa, numero=mesa_numero)

        return render(request, "cliente/mesa_qr.html", {
            "step": "pin",
            "mesa": mesa,
            "nombre": nombre,
            "personas": personas,
            "errores": [],
        })

    # -------------------------------
    # PASO 3 → CONFIRMACIÓN
    # -------------------------------
    if vista == "pin":

        mesa_numero = request.POST.get("mesa_numero")
        nombre = request.POST.get("nombre")
        personas = request.POST.get("personas")
        pin_ingresado = request.POST.get("pin")

        mesa = get_object_or_404(Mesa, numero=mesa_numero)

        errores = []
        if mesa.pin and pin_ingresado != mesa.pin:
            errores.append("PIN incorrecto.")
            return render(request, "cliente/mesa_qr.html", {
                "step": "pin",
                "mesa": mesa,
                "nombre": nombre,
                "personas": personas,
                "errores": errores
            })

        # Bloquear mesa
        mesa.estado = "ocupada"
        mesa.updated_at = timezone.now()
        mesa.save()

        # Pantalla final
        return render(request, "cliente/mesa_confirmada.html", {
            "mesa": mesa,
            "nombre": nombre,
            "personas": personas,
            "menu_url": reverse("menu"),
        })

    # 🔥 Si llega aquí, siempre retorna algo:
    return render(request, "cliente/mesa_qr.html", {
        "step": "form",
        "errores": ["Flujo inválido, volviendo al inicio."],
        "nombre": "",
        "personas": "",
    })


# @csrf_protect
# def confirmar_mesa(request, numero):
#     return redirect("mesa_qr", numero=numero)

# def confirmar_mesa_qr(request, mesa_id):
#     mesa = get_object_or_404(Mesa, id=mesa_id)
#     errores = []

#     if request.method == "POST":
#         nombre = request.POST.get("nombre", "").strip()
#         personas = request.POST.get("personas", "").strip()
#         pin_ingresado = request.POST.get("pin", "").strip()

#         # Validaciones
#         if not nombre:
#             errores.append("Debes ingresar tu nombre.")
#         if not personas.isdigit() or int(personas) < 1:
#             errores.append("Debes ingresar una cantidad válida de personas.")
#         if pin_ingresado != mesa.pin:
#             errores.append("El PIN ingresado no coincide con el de esta mesa.")

#         # Si todo está correcto → flujo final
#         if not errores:
#             return render(request, "mesa_confirmada.html", {
#                 "mesa": mesa,
#                 "nombre": nombre,
#                 "personas": personas,
#             })

#     return render(request, "mesa_qr.html", {
#         "mesa": mesa,
#         "mostrar_formulario": False,
#         "mostrar_mesas": False,
#         "errores": errores,
#         "pin_ingresado": request.POST.get("pin", ""),
#         "nombre": request.POST.get("nombre", ""),
#         "personas": request.POST.get("personas", "1"),
#     })


from django.shortcuts import redirect

# def seleccion_mesas(request):
#     """
#     Vista inicial del flujo cliente → formulario + mesas + PIN.
#     Redirige siempre a mesa 1 por defecto o muestra selección general.
#     """
#     # Mostrar la página base del flujo QR
#     return render(request, "cliente/mesa_qr.html", {})


# def seleccion_mesas(request):
#     mesas = Mesa.objects.all().order_by("numero")
#     return render(request, "cliente/seleccion_mesa.html", {"mesas": mesas})
